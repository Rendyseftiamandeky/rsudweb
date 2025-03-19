from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os, json
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'rahasia_super_aman'  # Ganti dengan secret key yang aman

# ----- Allowed Extensions & allowed_file -----
ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ----- Fungsi Bantuan User -----
def load_users():
    if not os.path.exists('users.json'):
        with open('users.json', 'w') as f:
            json.dump({}, f)
    try:
        with open('users.json', 'r') as f:
            return json.load(f)
    except json.JSONDecodeError:
        return {}

def save_users(users):
    with open('users.json', 'w') as f:
        json.dump(users, f, indent=4)

def valid_format(data):
    return all("-" in d and len(d.split("-")) == 3 for d in data)

# ----- Halaman Utama -----
@app.route('/')
def home():
    tanggal = datetime.now().strftime("%d-%m-%Y")
    return render_template('home.html', tanggal=tanggal)

# ----- Register -----
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        users = load_users()
        if username in users:
            flash('Username sudah terdaftar!')
            return redirect(url_for('register'))

        users[username] = {'password': password}
        save_users(users)
        flash('Berhasil mendaftar! Silakan login.')
        return redirect(url_for('login'))
    return render_template('register.html')

# ----- Login -----
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = load_users()
        if username in users and users[username]['password'] == password:
            session['username'] = username
            flash('Login berhasil!')
            return redirect(url_for('dashboard'))
        flash('Username atau password salah!')
        return redirect(url_for('login'))
    return render_template('login.html')

# ----- Dashboard Input Data -----
@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'username' not in session:
        flash('Silakan login dulu.')
        return redirect(url_for('login'))
    username = session['username']
    if request.method == 'POST':
        bulan = request.form['bulan']
        data_input = request.form['data'].strip().split()
        if not valid_format(data_input):
            flash('Format salah. Gunakan format 00-00-00 dan pisahkan dengan spasi.')
            return redirect(url_for('dashboard'))
        try:
            wb = load_workbook('contoh data.xlsx')
            ws = wb.active
            bulan_map = {
                'Januari': 'A', 'Februari': 'B', 'Maret': 'C', 'April': 'D',
                'Mei': 'E', 'Juni': 'F', 'Juli': 'G', 'Agustus': 'H',
                'September': 'I', 'Oktober': 'J', 'November': 'K', 'Desember': 'L'
            }
            col = bulan_map.get(bulan)
            if not col:
                flash('Bulan tidak valid.')
                return redirect(url_for('dashboard'))
            row = 2
            while ws[f"{col}{row}"].value:
                row += 1
            for item in data_input:
                ws[f"{col}{row}"] = item
                row += 1
            wb.save('contoh data.xlsx')
            flash('Data berhasil disimpan ke Excel.')
        except Exception as e:
            flash(f'Terjadi kesalahan: {e}')
    return render_template('dashboard.html', username=username)

# ----- Upload Template Excel -----
@app.route('/upload_template', methods=['GET', 'POST'])
def upload_template():
    if 'username' not in session:
        flash('Silakan login terlebih dahulu.')
        return redirect(url_for('login'))
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Tidak ada file yang diupload.')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('Tidak ada file yang dipilih.')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(os.getcwd(), 'contoh data.xlsx'))
            flash('Template Excel berhasil diperbarui.')
            return redirect(url_for('dashboard'))
        else:
            flash('Format file tidak diizinkan. Harus file .xlsx')
            return redirect(request.url)
    return render_template('upload_template.html')

# ----- Hapus Data -----
@app.route('/delete_data', methods=['GET', 'POST'])
def delete_data():
    if 'username' not in session:
        flash('Silakan login terlebih dahulu.')
        return redirect(url_for('login'))
    username = session['username']
    if request.method == 'POST':
        bulan = request.form['bulan']
        data_value = request.form['data_value'].strip()
        if not data_value:
            flash('Masukkan nilai data yang ingin dihapus.')
            return redirect(url_for('delete_data'))
        try:
            wb = load_workbook('contoh data.xlsx')
            ws = wb.active
            bulan_map = {
                'Januari': 'A', 'Februari': 'B', 'Maret': 'C', 'April': 'D',
                'Mei': 'E', 'Juni': 'F', 'Juli': 'G', 'Agustus': 'H',
                'September': 'I', 'Oktober': 'J', 'November': 'K', 'Desember': 'L'
            }
            col = bulan_map.get(bulan)
            if not col:
                flash('Bulan tidak valid.')
                return redirect(url_for('delete_data'))
            found = False
            row = 2
            while ws[f"{col}{row}"].value:
                if str(ws[f"{col}{row}"].value).strip() == data_value:
                    ws.delete_rows(row, 1)
                    found = True
                    break
                row += 1
            if found:
                wb.save('contoh data.xlsx')
                flash('Data berhasil dihapus.')
            else:
                flash('Data tidak ditemukan.')
        except Exception as e:
            flash(f'Terjadi kesalahan: {e}')
        return redirect(url_for('dashboard'))
    return render_template('delete_data.html', username=username)

# ----- Download Data -----
@app.route('/download')
def download():
    if 'username' not in session:
        flash('Silakan login terlebih dahulu.')
        return redirect(url_for('login'))
    try:
        return send_file('contoh data.xlsx', as_attachment=True)
    except Exception as e:
        flash(f'Terjadi kesalahan saat mengunduh file: {e}')
        return redirect(url_for('dashboard'))
    
    

# ----- Logout -----
@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout.')
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
